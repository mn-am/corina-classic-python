"""
This shows how CORINA Classic can be used as a file converter with the driver option -d no3d.
Furthermore, small fragments (e.g., counter ions in salts) should be removed and all molecules
should be neutralized using options -d rs,neu.

The input file is an excel file using the python package openpyxl
The output is a SDF file without 2D coordinates.

"""

import openpyxl
import corina

# load the workbook
wb = openpyxl.load_workbook("structure_clean_up_smiles_examples.xlsx")

# Get the sheet
sheet = wb.active

# Iterate to read the cell values
rows = []
for row in range(1, sheet.max_row):
    cell_values = []
    for col in sheet.iter_cols(1, sheet.max_column):
        cell_values.append(col[row].value)
    rows.append(cell_values)

# convert the data for CORINA SMILES input
rows_data = "\n".join([
    ("\t".join([str(cv) for cv in row])) for row in rows
])

# Create an instance of corinaBuffer
buffer = corina.CorinaBuffer()

# specify the input
buffer.input = rows_data

# specify run options
options = "-i t=smiles,sc#=2,nc#=1"  # find the SMILES correctly (SMILES in column 2 and name in column 1).
options += " -d no3d"  # no 3D computation
options += " -d rs,neu"  # remove small fragments and neutralize

# put together the command
buffer.command = "corina  " + options

# do the computation
buffer.proceed()

# print the generated structures in SD format
print(buffer.output)
